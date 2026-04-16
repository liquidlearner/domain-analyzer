import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Prioritized Services"
ws2 = wb.create_sheet("Summary by Team")

# ── Colors (matched to incident.io brand palette from original migration docs) ──
ORANGE    = "F25533"  # incident.io orange — header rows, section banners
CREAM     = "E2DACC"  # warm beige — pilot rows, summary section headers
HIGHLIGHT = "FFF8EC"  # warm cream — Sprint 0 / CET / action rows
HL_TEXT   = "D4700A"  # amber — text on highlight rows
ALT_ROW   = "F4F4F4"  # light grey — alternating data rows
WHITE     = "FFFFFF"
TEXT      = "2D2D2D"  # near-black body text
TEXT_GREY = "666666"  # secondary / metadata text
HEADER_FG = "FFFFFF"  # white text on orange headers

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Sheet 1: Prioritized Services ─────────────────────────────────────────────

col_widths = [12, 24, 36, 26, 10, 14, 8, 36, 40]
col_names  = ["Priority", "Sprint / Phase", "Service Name", "Team",
               "Status", "# Integrations", "CET?", "Escalation Policy", "Notes"]

for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Header row — orange background, white bold text
for col, name in enumerate(col_names, 1):
    cell = ws1.cell(row=1, column=col, value=name)
    cell.font = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill = fill(ORANGE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws1.row_dimensions[1].height = 22


def write_section_banner(ws, row_num, label):
    """Full-width section banner (Sprint 0 / Sprint 1 / Sprint 2 headings)."""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = Font(bold=True, size=10, color=HEADER_FG)
    cell.fill = fill(ORANGE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 18


def write_group_header(ws, row_num, label):
    """Sub-group header (team name + service count)."""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = Font(bold=True, size=9, color=TEXT)
    cell.fill = fill(ALT_ROW)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 15


def write_data_row(ws, row_num, data, bg, is_cet=False, bold=False):
    """Write a single service row."""
    actual_bg = HIGHLIGHT if is_cet else bg
    text_color = HL_TEXT if (actual_bg == HIGHLIGHT) else TEXT
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill(actual_bg)
        cell.font = Font(size=9, bold=bold, color=text_color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row_num].height = 15


r = 2

# ── SPRINT 0 ──────────────────────────────────────────────────────────────────
write_section_banner(ws1, r, "⚑  SPRINT 0 — Resolve before any team migrates")
r += 1

write_group_header(ws1, r, "  — DBE On-Call —  1 service")
r += 1
write_data_row(ws1, r, ["⚑ Sprint 0", "Sprint 0 — CET Review", "DBE Critical Alerts", "DBE On-Call", "warning", "4", "✓ CET", "DBE Escalation", "CET review required Sprint 0"], HIGHLIGHT, is_cet=True, bold=True)
r += 1

write_group_header(ws1, r, "  — Security Team —  7 services  (CET in Sprint 0; remaining 6 in Sprint 2)")
r += 1
write_data_row(ws1, r, ["⚑ Sprint 0", "Sprint 0 — CET Review", "aws-security-cloudwatch-alerts", "Security Team", "active", "1", "✓ CET", "Security Escalation", "CET review required Sprint 0"], HIGHLIGHT, is_cet=True, bold=True)
r += 1

write_group_header(ws1, r, "  — Support On-Call —  3 services  (CET in Sprint 0; remaining 2 in Sprint 2)")
r += 1
write_data_row(ws1, r, ["⚑ Sprint 0", "Sprint 0 — CET Review", "Slack Webhook", "Support On-Call", "active", "1", "✓ CET", "Support Leadership Incident Escalation Policy", "CET review required Sprint 0"], HIGHLIGHT, is_cet=True, bold=True)
r += 1

write_group_header(ws1, r, "  — Techops —  43 services  (CET in Sprint 0; 42 services in Sprint 1 Pilot)")
r += 1
write_data_row(ws1, r, ["⚑ Sprint 0", "Sprint 0 — CET Review", "CloudWatch Alerts", "Techops", "active", "1", "✓ CET", "SRE Escalation", "CET review required Sprint 0"], HIGHLIGHT, is_cet=True, bold=True)
r += 1

write_group_header(ws1, r, "  — No team assigned —  15 services  ⚠ Must assign owner before migration")
r += 1
unassigned = [
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "CloudZero", "No team assigned", "critical", "1", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Costello deals-api", "No team assigned", "active", "2", None, "Salesloft Astrologers Escalation", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Debriefer", "No team assigned", "active", "1", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Dial a Dev", "No team assigned", "warning", "2", None, "Page a developer for help", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Drift in the box", "No team assigned", "active", "1", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Elasticsearch Forecasts", "No team assigned", "active", "1", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Engineering Leadership Management", "No team assigned", "active", "0", None, "Manager On-call Escalation", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Incidents without monitors", "No team assigned", "active", "0", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Melody", "No team assigned", "active", "3", None, "Dev Null Escalation", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Panopticon Demo", "No team assigned", "critical", "1", None, "Test Team Escalation", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Planner", "No team assigned", "active", "0", None, "Apptio-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Product Feedback Bot", "No team assigned", "active", "2", None, "Product Feedback Bot-ep", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Query Monitor Alarms", "No team assigned", "active", "1", None, "Melody Engineering Teams", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "Test Service", "No team assigned", "active", "0", None, "Dev Null Escalation", "Unassigned — team must be added before migration"],
    ["⚑ Sprint 0", "Sprint 0 Action — Assign Team", "devnull", "No team assigned", "active", "1", None, "Dev Null Escalation", "Unassigned — team must be added before migration"],
]
for svc in unassigned:
    write_data_row(ws1, r, svc, HIGHLIGHT)
    r += 1

# ── SPRINT 1 ──────────────────────────────────────────────────────────────────
write_section_banner(ws1, r, "★  SPRINT 1 — Techops Pilot  (Apr 21 – May 2)")
r += 1
write_group_header(ws1, r, "  — Techops —  42 services")
r += 1
techops_pilot = [
    "ArgoCD", "Cert Manager", "Cluster Autoscaler", "Conversion Webhook", "DNS Checker",
    "Datadog", "Datadog Agent", "Elasticsearch Canary QA", "Fluentd", "GCP StackDriver Alerts",
    "Harbor", "Infra Logging", "Ingress Nginx", "Istio", "Istio QA", "Kafka", "Kate API",
    "Keyscore", "Kube State Metrics", "Kubernetes", "Kubernetes API", "Kyverno",
    "Kyverno low urgency", "Legacy Redirector", "NAT Low Urgency", "Panopticon",
    "Pending Pods Watcher", "Pgbouncer", "PyPi Server", "Replicaset Watcher", "SRE",
    "SRE-low-priority", "SalesLoft Docker Registry", "Support Initiated",
    "Twilio Logs Middleware", "Vault", "Vault-canary", "Vertical Pod Autoscaler",
    "Zoomverify", "elasticsearch", "k8s Spot Termination Handler", "reloader",
]
for i, svc in enumerate(techops_pilot):
    row_bg = WHITE if i % 2 == 0 else ALT_ROW
    write_data_row(ws1, r, ["★ Pilot", "Sprint 1 — Pilot", svc, "Techops", "active", "varies", None, "SRE Escalation", None], CREAM)
    r += 1

# ── SPRINT 2 ──────────────────────────────────────────────────────────────────
write_section_banner(ws1, r, "2  SPRINT 2 — Full Cutover  (May 3 – 9)  ·  All remaining 18 teams")
r += 1

sprint2_teams = {
    "Conversation Intelligence": [
        ("Conversations API", "active", "2", "Conversations Escalation"),
        ("Noteninja Background Jobs", "active", "3", "Conversations Escalation"),
        ("Noteninja Calendar", "active", "2", "Conversations Escalation"),
        ("Noteninja Meeting Recorder", "active", "3", "Conversations Escalation"),
        ("Noteninja Transcription Endpoint", "active", "3", "Conversations Escalation"),
        ("Noteninja Web Api", "active", "3", "Conversations Escalation"),
        ("noteninja-elasticsearch", "active", "1", "Conversations Escalation"),
        ("searchservice-elasticsearch", "active", "1", "Conversations Escalation"),
    ],
    "Deals": [
        ("Aurora", "active", "3", "Deals Escalation"), ("Costello Workers", "active", "2", "Deals Escalation"),
        ("Costello Workers Monitor", "active", "2", "Deals Escalation"), ("Costello abbott", "active", "3", "Deals Escalation"),
        ("Costello bud", "active", "3", "Deals Escalation"), ("Costello buster", "active", "3", "Deals Escalation"),
        ("Costello elvis", "active", "3", "Deals Escalation"), ("Costello keystone", "active", "3", "Deals Escalation"),
        ("Costello lloyd", "active", "3", "Deals Escalation"), ("Costello lou", "active", "3", "Deals Escalation"),
        ("Costello odessa", "active", "3", "Deals Escalation"), ("Costello opp-refinery", "active", "3", "Deals Escalation"),
        ("Costello parton", "active", "3", "Deals Escalation"), ("Costello paul", "active", "3", "Deals Escalation"),
        ("Costello redirector", "active", "3", "Deals Escalation"), ("Costello roscoe", "active", "3", "Deals Escalation"),
        ("Costello schedulers", "active", "2", "Deals Escalation"), ("Costello web", "active", "3", "Deals Escalation"),
        ("Costello webhooks", "active", "2", "Deals Escalation"), ("Deals AI", "active", "1", "Deals Escalation"),
        ("Deals PubSub", "active", "1", "Deals Escalation"), ("Dolly", "active", "3", "Deals Escalation"),
        ("deals-elasticsearch", "active", "1", "Deals Escalation"), ("deals-workers", "active", "1", "Deals Escalation"),
    ],
    "Little Five Endpoints": [
        ("API Docsite", "active", "3", "Little Five Endpoints Escalation"), ("Audit-Logs-Report-Generator", "active", "1", "Little Five Endpoints Escalation"),
        ("Bulk Service", "active", "3", "Little Five Endpoints Escalation"), ("Bulk-Service", "active", "1", "Little Five Endpoints Escalation"),
        ("Doculoft", "active", "2", "Little Five Endpoints Escalation"), ("External-ID", "active", "2", "Little Five Endpoints Escalation"),
        ("Frontend Integration Registry - FIRE", "active", "2", "Little Five Endpoints Escalation"), ("Global API Gateway", "active", "3", "Little Five Endpoints Escalation"),
        ("LFE", "active", "2", "Little Five Endpoints Escalation"), ("LIFEI", "active", "3", "Little Five Endpoints Escalation"),
        ("Notification Service", "active", "3", "Little Five Endpoints Escalation"), ("Observability", "active", "2", "Little Five Endpoints Escalation"),
        ("Portal", "active", "3", "Little Five Endpoints Escalation"), ("Regional API Gateway", "active", "3", "Little Five Endpoints Escalation"),
        ("SalesLoft Pusher", "active", "2", "Little Five Endpoints Escalation"), ("Tenant Service", "active", "3", "Little Five Endpoints Escalation"),
        ("Webhook Service", "active", "3", "Little Five Endpoints Escalation"), ("janus", "active", "1", "Little Five Endpoints Escalation"),
        ("signal-service", "active", "2", "Little Five Endpoints Escalation"), ("specula", "active", "1", "Little Five Endpoints Escalation"),
    ],
    "Security Team": [
        ("BetterCloud Sev1 Alerts", "active", "1", "Security Escalation"), ("EDR Test", "active", "1", "Security Escalation"),
        ("JAMF Alerts", "active", "1", "Security Escalation"), ("SEV1 - Security - PagerDuty Alert", "active", "2", "Security Escalation"),
        ("Security - Threat Stack Email", "active", "1", "Security Escalation"), ("Sumo Logic Email Integration", "active", "1", "Security Escalation"),
    ],
    "Support On-Call": [
        ("Salesforce", "active", "1", "Support Escalation Process"), ("Zendesk", "active", "1", "Support Escalation Process"),
    ],
    "Analytics and Coaching": [
        ("Analytics API", "warning", "3", "Analytics and Coaching"), ("Cube-Analytics", "critical", "2", "Analytics and Coaching"),
        ("DataMob", "active", "2", "Analytics and Coaching"), ("Goals API", "active", "2", "Analytics and Coaching"),
        ("Model Registry", "active", "3", "Analytics and Coaching"), ("Seniority", "active", "3", "Analytics and Coaching"),
        ("analytics-ai-agent", "active", "1", "Analytics and Coaching"),
    ],
    "CCR": [
        ("CCR", "active", "2", "Cadence Clearwater Revival Escalation"), ("Dialer Backend", "active", "3", "Cadence Clearwater Revival Escalation"),
        ("Dialer Integrations", "active", "2", "Cadence Clearwater Revival Escalation"), ("Genesys Wizard", "active", "2", "Cadence Clearwater Revival Escalation"),
        ("Keyscore Manager", "active", "3", "Cadence Clearwater Revival Escalation"), ("Mixmaster", "active", "3", "Cadence Clearwater Revival Escalation"),
    ],
    "Cloud9erz": [
        ("Dewey", "active", "3", "Cloud9erz Escalation"), ("Eos", "warning", "2", "Cloud9erz Escalation"),
        ("Gandalf", "active", "3", "Cloud9erz Escalation"), ("Gandalf Router", "active", "2", "Cloud9erz Escalation"),
        ("Houdini", "active", "2", "Cloud9erz Escalation"), ("Man In The Mirror Back", "active", "3", "Cloud9erz Escalation"),
        ("Migrater", "active", "3", "Cloud9erz Escalation"), ("Multipass", "active", "3", "Cloud9erz Escalation"),
        ("Picard", "active", "1", "Cloud9erz Escalation"), ("Search Service", "active", "3", "Cloud9erz Escalation"),
        ("Switchboard", "active", "3", "Cloud9erz Escalation"),
    ],
    "HSA": [
        ("Cerebro", "active", "3", "Heavy Swearer's Anonymous Escalation"), ("Email Gateway", "active", "3", "Heavy Swearer's Anonymous Escalation"),
        ("HSA", "active", "2", "Heavy Swearer's Anonymous Escalation"), ("Mail Spy", "active", "3", "Heavy Swearer's Anonymous Escalation"),
        ("Pathfinder", "active", "3", "Heavy Swearer's Anonymous Escalation"), ("Scintillas", "active", "3", "Heavy Swearer's Anonymous Escalation"),
        ("Scout", "active", "3", "Heavy Swearer's Anonymous Escalation"), ("Tenant Redirector", "active", "3", "Heavy Swearer's Anonymous Escalation"),
    ],
    "Lead IntelliAgent": [
        ("Chorus", "active", "1", "AKL Escalation Policy"),
    ],
    "No Data Left Behind": [
        ("CRM Repo", "active", "3", "No Data Left Behind Escalation"), ("CRM Security", "active", "3", "No Data Left Behind Escalation"),
        ("Melody - CRM, Sync, Default Queues", "active", "3", "No Data Left Behind Escalation"),
        ("Object-service", "active", "1", "No Data Left Behind Escalation"), ("Saasy", "active", "2", "No Data Left Behind Escalation"),
        ("Thanos", "active", "3", "No Data Left Behind Escalation"), ("XX Proxy", "active", "3", "No Data Left Behind Escalation"),
    ],
    "OMG": [
        ("Harmonize", "active", "3", "OMG Primary Escalation Policy"), ("Meetings API", "active", "3", "OMG Primary Escalation Policy"),
        ("Meetings Inspector", "active", "2", "OMG Primary Escalation Policy"), ("Meetings Router", "active", "3", "OMG Primary Escalation Policy"),
        ("Meetings UI", "active", "3", "OMG Primary Escalation Policy"), ("Outlook Connect", "active", "3", "OMG Primary Escalation Policy"),
    ],
    "Prospector Pod": [
        ("Global Web Research", "active", "1", "Prospector Escalation Policy"), ("Sherlock", "active", "1", "Prospector Escalation Policy"),
        ("lead-identification", "active", "1", "Prospector Escalation Policy"), ("zoominfo-integration", "active", "1", "Prospector Escalation Policy"),
    ],
    "Release Bot Support": [
        ("Release Bot", "warning", "1", "Release bot support escalation policy"),
    ],
    "RnB": [
        ("Rhapsody", "active", "3", "Rhythm and Blues Escalation"),
    ],
    "Wild Wild Data": [
        ("Ask Salesloft Anything Core", "active", "1", "Wild Wild Data Escalation"), ("Buyer Engagement Score", "active", "2", "Wild Wild Data Escalation"),
        ("Calculon", "active", "2", "Wild Wild Data Escalation"), ("Coaching Agent", "active", "1", "Wild Wild Data Escalation"),
        ("Deal Score", "active", "2", "Wild Wild Data Escalation"), ("Deals Chatbot", "active", "2", "Wild Wild Data Escalation"),
        ("Email Science", "active", "1", "Wild Wild Data Escalation"), ("Engagement Score", "active", "3", "Wild Wild Data Escalation"),
        ("Forecast AI", "active", "2", "Wild Wild Data Escalation"), ("Job Function", "active", "2", "Wild Wild Data Escalation"),
        ("LangFuse", "active", "1", "Wild Wild Data Escalation"), ("LiteLLM", "active", "3", "Wild Wild Data Escalation"),
        ("Mario", "active", "2", "Wild Wild Data Escalation"),
    ],
    "Workflow Pod": [
        ("Activity Service", "active", "3", "Workflow Escalation"), ("Scoreboard", "active", "3", "Workflow Escalation"),
        ("Signals Processor", "active", "2", "Workflow Escalation"), ("WUT", "active", "2", "WUT Escalation"),
        ("Workflow Melody", "active", "2", "Workflow Escalation"),
    ],
}

for team, services in sprint2_teams.items():
    write_group_header(ws1, r, f"  — {team} —  {len(services)} service{'s' if len(services) != 1 else ''}")
    r += 1
    for i, svc in enumerate(services):
        row_bg = WHITE if i % 2 == 0 else ALT_ROW
        write_data_row(ws1, r, ["Sprint 2", "Sprint 2", svc[0], team, svc[1], svc[2], None, svc[3], None], row_bg)
        r += 1

ws1.freeze_panes = "A2"

# ── Sheet 2: Summary by Team ───────────────────────────────────────────────────

summary_headers = ["Team", "Services", "Sprint", "CETs", "Unassigned?"]
for col, name in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=name)
    cell.font = Font(bold=True, color=HEADER_FG, size=10)
    cell.fill = fill(ORANGE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 26
ws2.column_dimensions["D"].width = 8
ws2.column_dimensions["E"].width = 12

sr = 2


def sum_section(ws, row, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, size=10, color=TEXT)
    cell.fill = fill(CREAM)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def sum_row(ws, row, team, svcs, sprint, cets, unassigned, bg=WHITE):
    text_color = HL_TEXT if bg == HIGHLIGHT else TEXT
    data = [team, svcs, sprint, cets, unassigned]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill(bg)
        cell.font = Font(size=9, color=text_color)
        cell.alignment = Alignment(vertical="center")
        cell.border = border


sum_section(ws2, sr, "Sprint 0 — CET Review / Team Assignment")
sr += 1
sum_row(ws2, sr, "DBE On-Call", 1, "Sprint 0", 1, None, HIGHLIGHT); sr += 1
sum_row(ws2, sr, "No team assigned", 15, "Sprint 0", 0, "⚠ Yes", HIGHLIGHT); sr += 1
sum_row(ws2, sr, "Security Team", 7, "Sprint 0 (CET) + Sprint 2", 1, None, HIGHLIGHT); sr += 1
sum_row(ws2, sr, "Support On-Call", 3, "Sprint 0 (CET) + Sprint 2", 1, None, HIGHLIGHT); sr += 1
sum_row(ws2, sr, "Techops", 43, "Sprint 0 (CET) + Sprint 1", 1, None, HIGHLIGHT); sr += 1

sr += 1
sum_section(ws2, sr, "Sprint 1 — Techops Pilot  (Apr 21 – May 2)")
sr += 1
sum_row(ws2, sr, "Techops", 42, "Sprint 1 — Pilot", 0, None, CREAM); sr += 1

sr += 1
sum_section(ws2, sr, "Sprint 2 — Full Cutover  (May 3 – 9)")
sr += 1
sprint2_summary = [
    ("Conversation Intelligence", 8, 0), ("Deals", 24, 0),
    ("Little Five Endpoints", 20, 0), ("Security Team", 6, 1),
    ("Support On-Call", 2, 1), ("Analytics and Coaching", 7, 0),
    ("CCR", 6, 0), ("Cloud9erz", 11, 0), ("HSA", 8, 0),
    ("Lead IntelliAgent", 1, 0), ("No Data Left Behind", 7, 0),
    ("OMG", 6, 0), ("Prospector Pod", 4, 0), ("Release Bot Support", 1, 0),
    ("RnB", 1, 0), ("Wild Wild Data", 13, 0), ("Workflow Pod", 5, 0),
]
for i, (team, svcs, cets) in enumerate(sprint2_summary):
    row_bg = WHITE if i % 2 == 0 else ALT_ROW
    sum_row(ws2, sr, team, svcs, "Sprint 2", cets, None, row_bg)
    sr += 1

sr += 1
# Total row
for col, val in enumerate(["TOTAL", 191, None, 4, None], 1):
    cell = ws2.cell(row=sr, column=col, value=val)
    cell.font = Font(bold=True, size=10, color=HEADER_FG)
    cell.fill = fill(ORANGE)
    cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
    cell.border = border

ws2.freeze_panes = "A2"

out_path = "/sessions/friendly-vigilant-ride/mnt/domain-analyzer/salesloft-prioritized-services-2026-04-13.xlsx"
wb.save(out_path)
print("Saved:", out_path)
